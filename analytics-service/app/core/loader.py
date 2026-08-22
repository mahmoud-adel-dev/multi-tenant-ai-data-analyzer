"""Secure file loading with hard limits (zip-bomb, row/column caps)."""
from __future__ import annotations

import csv
import io
import json
from dataclasses import dataclass, field
from typing import Any

import openpyxl
import polars as pl

from app.core.config import settings
from app.core.exceptions import MalformedFileError, UnsupportedFileError


@dataclass
class LoadedTable:
    frame: pl.DataFrame
    detected_delimiter: str | None = None
    encoding: str = "utf-8"
    warnings: list[str] = field(default_factory=list)


MAX_SHEETS = 50
MAX_EXCEL_CELLS = 5_000_000


def _enforce_frame_limits(lf: pl.LazyFrame) -> pl.DataFrame:
    """Collects a lazy frame while enforcing row/column ceilings."""
    width = lf.collect_schema().len()
    if width > settings.max_columns:
        raise MalformedFileError(f"Dataset has {width} columns; maximum allowed is {settings.max_columns}.")
    df = lf.head(settings.max_rows + 1).collect()
    if df.height > settings.max_rows:
        df = df.head(settings.max_rows)
        # Caller may warn about truncation via df.height check.
    return df


def _sniff_delimiter(sample: str) -> str | None:
    for candidate in [",", "\t", ";", "|"]:
        if sample.count(candidate) >= sample.count("\n"):
            return candidate
    return None


def load_table(buffer: bytes, file_type: str | None) -> LoadedTable:
    """Loads CSV/TSV/JSON/XLSX/XLS bytes into a Polars DataFrame with limits."""
    if len(buffer) == 0:
        raise MalformedFileError("File is empty.")

    ftype = (file_type or "").lower()

    if ftype in {"xlsx", "xls"} or (not ftype and buffer[:4] == b"PK\x03\x04"):
        return _load_excel(buffer)
    if ftype == "tsv":
        return _load_csv(buffer, delimiter="\t")
    if ftype == "csv":
        delim = _sniff_delimiter(buffer[:8192].decode("utf-8", errors="replace"))
        return _load_csv(buffer, delimiter=delim or ",")
    if ftype == "json":
        return _load_json(buffer)

    # Unknown type — try structural inference.
    if buffer[:4] == b"PK\x03\x04":
        return _load_excel(buffer)
    try:
        text = buffer[:4096].decode("utf-8")
        if not text:
            raise UnsupportedFileError("Unsupported or empty file content.")
        delim = _sniff_delimiter(text)
        return _load_csv(buffer, delimiter=delim or ",")
    except UnicodeDecodeError as exc:
        raise UnsupportedFileError("Unsupported file format.") from exc


def _load_csv(buffer: bytes, delimiter: str) -> LoadedTable:
    encoding = "utf-8"
    try:
        text = buffer.decode("utf-8")
    except UnicodeDecodeError:
        text = buffer.decode("latin-1")
        encoding = "latin-1"

    if not text.strip():
        raise MalformedFileError("File contains no data rows.")

    # Count rows cheaply to enforce the cap before parsing. Trailing newlines
    # would inflate the count and trigger a false truncation warning.
    line_count = text.rstrip("\r\n").count("\n") + 1
    truncated = False
    if line_count > settings.max_rows + 1_000_000:
        raise MalformedFileError(f"Dataset exceeds the maximum of {settings.max_rows:,} rows.")

    try:
        dialect = csv.Sniffer().sniff(text[:16384], delimiters=",\t;|")
        delimiter = dialect.delimiter
    except csv.Error:
        pass

    lf = pl.scan_csv(
        io.StringIO(text),
        separator=delimiter,
        infer_schema_length=min(10_000, max(1, line_count)),
        truncate_ragged_lines=False,
        try_parse_dates=True,
    )
    df = _enforce_frame_limits(lf)
    warnings: list[str] = []
    if df.height < line_count - 1:
        warnings.append(f"Dataset truncated to {settings.max_rows:,} rows for analysis.")
    return LoadedTable(frame=df, detected_delimiter=delimiter, encoding=encoding, warnings=warnings)


def _load_excel(buffer: bytes) -> LoadedTable:
    """Excel loading guarded against zip bombs: sheet/cell ceilings enforced."""
    import zipfile

    if not zipfile.is_zipfile(io.BytesIO(buffer)):
        # Legacy .xls is not ZIP-based; openpyxl cannot read it.
        raise UnsupportedFileError(
            "Legacy .xls workbooks are not supported. Please re-save as .xlsx or CSV."
        )

    wb = openpyxl.load_workbook(
        io.BytesIO(buffer),
        read_only=True,
        data_only=True,
        keep_links=False,
    )
    try:
        if len(wb.sheetnames) > MAX_SHEETS:
            raise MalformedFileError(f"Workbook has more than {MAX_SHEETS} sheets.")
        sheet = wb[wb.sheetnames[0]]
        total_cells = 0
        rows: list[list[Any]] = []
        header: list[str] | None = None
        for row in sheet.iter_rows(values_only=True):
            total_cells += sum(1 for v in row if v is not None)
            if total_cells > MAX_EXCEL_CELLS:
                raise MalformedFileError(f"Workbook exceeds {MAX_EXCEL_CELLS:,} cells.")
            if header is None:
                if all(v is None for v in row):
                    continue
                header = [str(v) if v is not None else f"column_{i}" for i, v in enumerate(row)]
                continue
            rows.append(list(row))
            if len(rows) > settings.max_rows:
                break
    finally:
        wb.close()

    if header is None:
        raise MalformedFileError("Workbook's first sheet has no header row.")
    # Only the first (primary) sheet is analyzed; additional sheets are ignored.
    df = pl.DataFrame(rows, schema=header, orient="row", strict=False)
    return LoadedTable(
        frame=df,
        warnings=[f"{len(wb.sheetnames) - 1} additional sheet(s) were not analyzed."],
    )


def _load_json(buffer: bytes) -> LoadedTable:
    # ``utf-8-sig`` accepts normal UTF-8 and removes a leading BOM when present.
    text = buffer.decode("utf-8-sig", errors="replace")
    try:
        data = json.loads(text)
    except json.JSONDecodeError as exc:
        raise MalformedFileError("Invalid JSON.") from exc

    records: list[Any]
    if isinstance(data, list):
        records = data
    elif isinstance(data, dict):
        # Prefer the largest list-valued key (common export pattern).
        list_keys = [(k, v) for k, v in data.items() if isinstance(v, list)]
        if list_keys:
            key, values = max(list_keys, key=lambda kv: len(kv[1]))
            records = values
        else:
            records = [data]
    else:
        raise MalformedFileError("JSON must be an array or object of records.")

    if not records:
        raise MalformedFileError("JSON array contains no records.")

    df = pl.DataFrame(records, strict=False)
    return _finish_json(df)


def _finish_json(df: pl.DataFrame) -> LoadedTable:
    if df.width > settings.max_columns:
        raise MalformedFileError(f"Dataset has {df.width} columns; maximum allowed is {settings.max_columns}.")
    if df.height > settings.max_rows:
        df = df.head(settings.max_rows)
    # Normalize all columns to strings-safe representations where needed.
    return LoadedTable(frame=df)
